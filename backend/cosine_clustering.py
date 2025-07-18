import pandas as pd
from sklearn.metrics.pairwise import cosine_similarity
import unicodedata
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


from data_cleaning import standardize_value
def cluster_column(df, column_name, threshold=0.8):
    print("Starting clustering")
    print("Column name:", column_name)
    print("Threshold:", threshold)

    try:
        df[column_name] = df[column_name].astype(str)
        df['Processed'] = df[column_name].apply(standardize_value)
        print("Sample processed values:", df['Processed'].head().tolist())

        vectorizer = TfidfVectorizer()
        x = vectorizer.fit_transform(df['Processed'])
        print("TF-IDF matrix created. Shape:", x.shape)

        sim_matrix = cosine_similarity(x)
        print("Cosine similarity matrix shape:", sim_matrix.shape)

        updated_column = df[column_name].copy()
        already_asked = set()

        for i in range(len(df)):
            for j in range(i + 1, len(df)):
                sim = sim_matrix[i][j]
                if sim >= threshold and updated_column[i] != updated_column[j]:
                    pair = (updated_column[i], updated_column[j])
                    if pair in already_asked:
                        continue
                    already_asked.add(pair)
                    df.at[j, column_name] = df.at[i, column_name]

        #df[f'Updated_{column_name}'] = updated_column
        #print("Clustering completed")
        #return df[[column_name, f'Updated_{column_name}']]
        df.drop(columns=['Processed'], inplace=True)
        return df

    except Exception as e:
        print("ERROR in cluster_column:", str(e))
        raise

""" def highlight_changes_in_excel(csv_path, column_name, output_excel_path):
    df = pd.read_csv(csv_path)

    # Save to Excel
    df.to_excel(output_excel_path, index=False)

    # Load workbook to modify formatting
    wb = load_workbook(output_excel_path)
    ws = wb.active

    changed_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow

    # Loop through rows and highlight changed rows
    for i, row in enumerate(df.itertuples(), start=2):  # Excel rows start at 1, header is at row 1
        if getattr(row, column_name) != getattr(row, f'Updated_{column_name}'):
            for cell in ws[i]:
                cell.fill = changed_fill

    wb.save(output_excel_path)
    print(f"Highlighted Excel saved to {output_excel_path}") """

#changed code below for 3 filter category

def highlight_changes_in_excel(original_df, updated_df, column_name, output_excel_path):
    # Save updated data to Excel
    updated_df.to_excel(output_excel_path, index=False)

    wb = load_workbook(output_excel_path)
    ws = wb.active

    changed_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for i in range(len(original_df)):
        if original_df.at[i, column_name] != updated_df.at[i, column_name]:
            for cell in ws[i + 2]:  # +2 because Excel rows are 1-indexed, row 1 is header
                cell.fill = changed_fill

    wb.save(output_excel_path)
    print(f"🟨 Highlighted Excel saved to {output_excel_path}")


""" 
def get_replacement_suggestions(df, column_name, threshold=0.8):
    df[column_name] = df[column_name].astype(str)
    df['Processed'] = df[column_name].apply(standardize_value)

    vectorizer = TfidfVectorizer()
    x = vectorizer.fit_transform(df['Processed'])

    sim_matrix = cosine_similarity(x)
    suggestions = []

    for i in range(len(df)):
        for j in range(i + 1, len(df)):
            sim = sim_matrix[i][j]
            if df['Processed'][i] == df['Processed'][j] or sim == 1.0:
                print(f"Skipping exact match: Row {i} and Row {j}, sim = {sim}")
                continue
            
            if sim >= threshold and df[column_name][i] != df[column_name][j]:
                
                
                suggestions.append({
                    "replace": {
                        "row": j,
                        "original": df[column_name][j],
                        "suggested_with_row": i,
                        "suggested_value": df[column_name][i],
                        "similarity": round(sim, 2)
                    }
                })
    return suggestions
 """


def get_replacement_suggestions(df, column_name, threshold=0.8):
    df[column_name] = df[column_name].astype(str)
    df['Processed'] = df[column_name].apply(standardize_value)

    vectorizer = TfidfVectorizer()
    x = vectorizer.fit_transform(df['Processed'])

    sim_matrix = cosine_similarity(x)
    suggestions = []

    for i in range(len(df)):
        for j in range(i + 1, len(df)):
            sim = sim_matrix[i][j]
            
            # Skip if similarity is 1.0 (100% match) or if they're already identical
            if sim == 1.0 or df[column_name][i] == df[column_name][j]:
                print(f"Skipping exact match: Row {i} and Row {j}, sim = {sim}")
                continue
            
            # Only suggest if similarity is above threshold but not 100%
            if sim >= threshold and sim < 1.0:
                suggestions.append({
                    "replace": {
                        "row": j,
                        "original": df[column_name][j],
                        "suggested_with_row": i,
                        "suggested_value": df[column_name][i],
                        "similarity": round(sim, 2)
                    }
                })
    return suggestions